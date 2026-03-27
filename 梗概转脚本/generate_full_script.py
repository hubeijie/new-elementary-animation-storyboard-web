#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
故事梗概 + 教研动效说明 xlsx → 完整五列脚本 xlsx。
支持 Gemini / OpenAI / Anthropic Claude / OpenAI 兼容（国内 DeepSeek、通义兼容模式等），见 --provider 与环境变量。
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


def read_synopsis(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def extract_teaching_rows(xlsx_path: Path) -> tuple[str, list[dict]]:
    """选取最匹配的工作表，返回 (表名, [{画面,逐字稿,动效,备注}, ...])"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    def score_header(row: tuple) -> tuple[dict[str, int], int]:
        cells = [str(c or "").replace("\n", " ").strip() for c in row]

        def find_col(aliases: list[str]) -> int:
            for i, c in enumerate(cells):
                for a in aliases:
                    if c == a or a in c:
                        return i
            return -1

        idx = {
            "画面": find_col(["画面", "环节", "模块"]),
            "逐字稿": find_col(["逐字稿", "台词"]),
            "动效": find_col(["动效", "动画"]),
            "备注": find_col(["备注", "说明"]),
        }
        s = 0
        if idx["逐字稿"] >= 0:
            s += 3
        if idx["动效"] >= 0:
            s += 3
        if idx["画面"] >= 0:
            s += 1
        return idx, s

    best = None
    best_score = -1
    best_off = 0
    best_name = ""

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        for off in range(min(3, len(rows))):
            idx, sc = score_header(rows[off])
            if sc > best_score:
                best_score = sc
                best = rows
                best_off = off
                best_name = name

    wb.close()

    if best_score < 4:
        raise SystemExit("未找到含「逐字稿」与「动效」列的表头（前 3 行内）。")

    header_idx, _ = score_header(best[best_off])
    idx = header_idx
    body = best[best_off + 1 :]
    out: list[dict] = []
    for row in body:
        if not row:
            continue

        def get(ci: int) -> str:
            if ci < 0 or ci >= len(row):
                return ""
            v = row[ci]
            return str(v).strip() if v is not None else ""

        item = {
            "画面": get(idx["画面"]),
            "逐字稿": get(idx["逐字稿"]),
            "动效": get(idx["动效"]),
            "备注": get(idx["备注"]),
        }
        if any(item.values()):
            out.append(item)

    return best_name, out


def build_prompt(title: str, synopsis: str, teaching: list[dict]) -> str:
    spec = """【脚本规范 · elementary-animation-storyboard-script · 与 Cursor Skill 同步】
1）输出一个 Markdown 表格，五列且顺序固定：镜号 | 画面 | 台词 | 画面描述 | 备注。
2）第一行表头必须是：| 镜号 | 画面 | 台词 | 画面描述 | 备注 |
3）第二行必须是分隔行：| --- | --- | --- | --- | --- |
4）台词与画面描述必须用相同编号 【1】【2】【3】… 逐步对齐；若教研逐字稿含 【0】或与 【1】混用，一律照抄保留。
5）【教研台词保真】教研 JSON 每一行的「逐字稿」字段：对应教学镜的「台词」列必须 **逐字拷贝**（含换行与标点），不得改写、不得调换句序、不得与另一行逐字稿合并成一句。无逐字稿而仅有画面/动效的行按表意单独成镜时在备注说明。
6）【剧情仅衔接】故事梗概拆成的角色对白、旁白、转场：只允许作为「衔接」出现在某一教学块之前、之后或两教学块之间；可单独成镜；备注可标「衔接」。禁止把衔接句与教研原句揉成一句导致教研被改。
7）【画面描述 · 分行体】将「动效」写入「画面描述」时：每个 【n】 单独占一行（Markdown 单元格内用 <br> 换行）；每条 *复用场景-xxx* *复用道具-xxx* *复用人设-xxx* 各占一行；可选第一行写镜头总述（如「切特写。」）。不得把多段 【n】 横排粘成一段。
8）故事梗概是剧情主体：按时间顺序拆成多镜；不要把梗概单独堆在表前几行当「参考区」。
9）「画面」列可作环节标签（如 引入、P1、衔接·xxx）。
10）除表格外不要输出任何解释文字。
"""
    return (
        spec
        + "\n【课题名】"
        + title
        + "\n\n【故事梗概 · 叙事主轴】\n"
        + synopsis
        + "\n\n【教研动效说明 · JSON】\n"
        + json.dumps(teaching, ensure_ascii=False, indent=2)
        + "\n\n请现在只输出 Markdown 表格。"
    )


def strip_fence(text: str) -> str:
    text = text.strip()
    m = re.search(r"```(?:markdown|md)?\s*([\s\S]*?)```", text, re.I)
    if m:
        return m.group(1).strip()
    return text


def parse_md_table(text: str) -> list[list[str]]:
    raw = strip_fence(text)
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    start = -1
    for i, ln in enumerate(lines):
        if re.search(r"\|?\s*镜号\s*\|", ln):
            start = i
            break
    if start < 0:
        raise ValueError("未找到以「镜号」开头的表头行")

    rows: list[list[str]] = []
    for ln in lines[start + 1 :]:
        if not ln.startswith("|"):
            break
        if re.match(r"^\|[\s\-:|]+\|?$", ln):
            continue
        parts = ln.split("|")
        if len(parts) < 2:
            break
        end = len(parts) - 1 if parts[-1] == "" else len(parts)
        cells = [re.sub(r"<br\s*/?>", "\n", c, flags=re.I).strip() for c in parts[1:end]]
        if cells and "镜号" in cells[0]:
            continue
        while len(cells) < 5:
            cells.append("")
        rows.append(cells[:5])
        if not any(rows[-1]):
            rows.pop()
    if not rows:
        raise ValueError("解析后无数据行")
    return rows


def call_gemini(api_key: str, model: str, prompt: str) -> str:
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    body = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.35, "maxOutputTokens": 8192},
    }
    with httpx.Client(timeout=180) as client:
        r = client.post(url, params={"key": api_key}, json=body)
        r.raise_for_status()
        data = r.json()
    try:
        return data["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError) as e:
        raise RuntimeError(f"Gemini 返回异常: {data}") from e


def call_openai(api_key: str, model: str, prompt: str) -> str:
    return call_openai_compatible(
        api_key, model, prompt, "https://api.openai.com/v1"
    )


def call_openai_compatible(api_key: str, model: str, prompt: str, base_url: str) -> str:
    base = base_url.rstrip("/")
    url = f"{base}/chat/completions"
    body = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.35,
        "max_tokens": 8192,
    }
    headers = {"Authorization": f"Bearer {api_key}"}
    with httpx.Client(timeout=180) as client:
        r = client.post(url, json=body, headers=headers)
        data = r.json()
    if r.is_error:
        err = data.get("error", {}) if isinstance(data, dict) else {}
        msg = err.get("message", r.text)
        raise RuntimeError(f"OpenAI 兼容接口错误 {r.status_code}: {msg}")
    try:
        return data["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as e:
        raise RuntimeError(f"OpenAI 兼容接口返回异常: {data}") from e


def call_claude(api_key: str, model: str, prompt: str) -> str:
    url = "https://api.anthropic.com/v1/messages"
    body = {
        "model": model,
        "max_tokens": 8192,
        "messages": [{"role": "user", "content": prompt}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
    }
    with httpx.Client(timeout=180) as client:
        r = client.post(url, json=body, headers=headers)
        data = r.json()
    if r.is_error:
        err = data.get("error", {}) if isinstance(data, dict) else {}
        msg = err.get("message", r.text)
        raise RuntimeError(f"Claude 错误 {r.status_code}: {msg}")
    try:
        block = data["content"][0]
        if block.get("type") == "text":
            return block["text"]
        raise KeyError("text")
    except (KeyError, IndexError, TypeError) as e:
        raise RuntimeError(f"Claude 返回异常: {data}") from e


def write_output_xlsx(
    out_path: Path,
    script_rows: list[list[str]],
    source_xlsx: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "脚本表·完整剧情"
    ws.append(["镜号", "画面", "台词", "画面描述", "备注"])
    for row in script_rows:
        r = (row + ["", "", "", "", ""])[:5]
        ws.append(r)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = wrap_top

    src = load_workbook(source_xlsx, read_only=True, data_only=True)
    for name in src.sheetnames:
        wsn = ("教研·" + name)[:31]
        wsd = wb.create_sheet(title=wsn)
        sh = src[name]
        for r in sh.iter_rows(values_only=True):
            wsd.append(["" if c is None else c for c in r])
    src.close()
    wb.save(out_path)


def main() -> None:
    p = argparse.ArgumentParser(
        description="梗概 + 教研 xlsx → 完整脚本 xlsx（Gemini / OpenAI / Claude / OpenAI 兼容）"
    )
    p.add_argument("--synopsis", type=Path, required=True, help="故事梗概文本文件 utf-8")
    p.add_argument("--teaching", type=Path, required=True, help="教研动效说明 .xlsx")
    p.add_argument("--out", type=Path, default=None, help="输出 .xlsx")
    p.add_argument("--title", type=str, default="", help="课题名")
    p.add_argument(
        "--provider",
        choices=("gemini", "openai", "claude", "openai_compat"),
        default="gemini",
        help="API 服务商；openai_compat 为 OpenAI 兼容（国内 DeepSeek、通义兼容模式等）",
    )
    p.add_argument(
        "--base-url",
        type=str,
        default="",
        help="OpenAI 兼容接口 Base URL（到 …/v1 或文档要求的路径，不含 /chat/completions）",
    )
    p.add_argument(
        "--model",
        type=str,
        default="",
        help="模型名；省略则用各服务商默认模型",
    )
    args = p.parse_args()

    defaults = {
        "gemini": "gemini-2.0-flash",
        "openai": "gpt-4o-mini",
        "claude": "claude-sonnet-4-20250514",
        "openai_compat": "deepseek-chat",
    }
    model = args.model.strip() or defaults[args.provider]

    env_keys = {
        "gemini": "GEMINI_API_KEY",
        "openai": "OPENAI_API_KEY",
        "claude": "ANTHROPIC_API_KEY",
        "openai_compat": "OPENAI_COMPAT_API_KEY",
    }
    api_key = os.environ.get(env_keys[args.provider], "").strip()
    if args.provider == "openai_compat" and not api_key:
        api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        hint = env_keys[args.provider]
        if args.provider == "openai_compat":
            hint += "（或未设置时可用 OPENAI_API_KEY）"
        print(f"请设置环境变量 {hint}", file=sys.stderr)
        sys.exit(1)

    synopsis = read_synopsis(args.synopsis)
    if not synopsis:
        print("梗概文件为空", file=sys.stderr)
        sys.exit(1)

    sheet_name, teaching = extract_teaching_rows(args.teaching)
    print(f"教研主表: {sheet_name}，共 {len(teaching)} 行", file=sys.stderr)

    title = args.title or args.teaching.stem
    prompt = build_prompt(title, synopsis, teaching)
    if args.provider == "openai_compat":
        base_url = (args.base_url or os.environ.get("OPENAI_COMPAT_BASE_URL", "")).strip()
        if not base_url:
            print(
                "openai_compat 须提供 --base-url 或环境变量 OPENAI_COMPAT_BASE_URL",
                file=sys.stderr,
            )
            sys.exit(1)
    print(f"调用 {args.provider}（{model}）…", file=sys.stderr)
    if args.provider == "gemini":
        text = call_gemini(api_key, model, prompt)
    elif args.provider == "openai":
        text = call_openai(api_key, model, prompt)
    elif args.provider == "openai_compat":
        text = call_openai_compatible(api_key, model, prompt, base_url)
    else:
        text = call_claude(api_key, model, prompt)
    rows = parse_md_table(text)
    print(f"解析到 {len(rows)} 行脚本", file=sys.stderr)

    out = args.out or args.teaching.with_name(f"{title}_完整脚本.xlsx".replace("/", "_"))
    write_output_xlsx(out, rows, args.teaching)
    print(out)


if __name__ == "__main__":
    main()
